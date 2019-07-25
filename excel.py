import os
import sys

from openpyxl import load_workbook, Workbook


class ExcelFileReadOnly:

    def __init__(self, path):
        self.path = path

    def load_file(self):
        self.wb = load_workbook(
            filename=self.path, read_only=True, data_only=True)
        ws = self.wb.worksheets[0]
        self.ws = ws
        self.wsv = ws.values

    def read_file(self):
        ws = self.ws
        wsv = self.wsv
        models = []
        models_fields = []
        models_ids = []
        for row in ws.iter_rows(min_row=1, max_row=1):
            models_data = row
        for model in models_data:
            if model.value:
                model = str(model.value).strip()
                if model:
                    models.append(model)
        for n in range(len(models)):
            for row in ws.iter_rows(min_row=n+2, max_row=n+2):
                model_fields = []
                for cell in row[1:]:
                    if cell.value and str(cell.value).strip():
                        cell = str(cell.value).strip()
                        model_fields.append(cell)
                models_fields.append(model_fields)
                models_ids.append(str(row[0].value))
        self.models_fields = models_fields
        self.models = models
        self.models_ids = models_ids
        rows_data = []
        #sys.stderr.write("\nmodelcount:" + str(len(models)))
        for row in ws.iter_rows(min_row=2+len(models)):
            # needs to check if it is not integer and raise error in that case.
            model_id = row[0].value
            if not model_id:
                continue
            model_id = str(model_id)
            model_index = models_ids.index(model_id)
            model_name = models[model_index]
            model_fields = models_fields[model_index]
            cid = row[len(model_fields)+1].value
            if cid:
                cid = int(cid)
            else:
                cid = None
            row_data = {"row": row[0].row, "id": cid, "fields": {
            }, "model": model_name}  # row is 1 based
            for i in range(0, len(model_fields)):
                if row[i + 1].value:
                    row_data["fields"][model_fields[i]] = str(row[i + 1].value)
                else:
                    row_data["fields"][model_fields[i]] = None
            rows_data.append(row_data)
        # [{"row":int, "id":int, "fields":{"fieldName":str_val,}, "model": str_model_name}]
        return rows_data

    def close(self):
        self.wb.close()


class ExcelFile(ExcelFileReadOnly):

    def load_file(self):
        self.wb = load_workbook(filename=self.path)
        ws = self.wb.worksheets[0]
        self.ws = ws
        self.wsv = ws.values

    def set_id(self, row, fields, id):
        self.ws.cell(row, len(fields)+2).value = id
        #cell = self.ws["A" + str(row_num)]
        #cell.value = id

    def create_file(self):
        self.wb = Workbook()
        self.ws = self.wb.worksheets[0]
        self.ws.title = "Anki Cards"

    def write(self, notes, models, col_width):
        ws = self.ws
        first_line = []
        headers = []
        for model in models:
            first_line.append(model["name"])
            hd = [model["id"]]
            hd += model["flds"]
            headers.append(hd)
        # write headers
        for n in range(len(first_line)):
            ws.cell(row=1, column=n + 1).value = first_line[n]
        for n in range(len(headers)):
            for m in range(len(headers[n])):
                ws.cell(row=n + 2, column=m + 1).value = headers[n][m]
        # write notes
        crow = len(headers) + 1  # 1 row before first row of note rows
        for note in notes:
            crow += 1
            model = note.model()
            val_row = [None]*len(model["flds"])
            for mdl in models:
                if mdl["name"] == model["name"]:
                    thismodel = mdl
                    break
            ws.cell(row=crow, column=1).value = str(thismodel["id"])
            for n in range(len(model["flds"])):
                for m in range(len(thismodel["flds"])):
                    if thismodel["flds"][m] == model["flds"][n]["name"]:
                        val_row[m] = note.fields[n]
                        break
            for n in range(len(val_row)):
                ws.cell(row=crow, column=n+2).value = str(val_row[n])
            ws.cell(row=crow, column=len(model["flds"]) + 2).value = note.id
        for x in range(len(col_width)):
            cl = ws.cell(row=1, column=x+1)
            if cl:
                col = cl.column_letter
                ws.column_dimensions[col].width = col_width[x]

    def save(self):
        dir = os.path.dirname(self.path)
        if not os.path.exists(dir):
            os.makedirs(dir)
        self.wb.save(filename=self.path)
